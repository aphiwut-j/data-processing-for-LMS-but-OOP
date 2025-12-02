import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class TableConverter:
    def __init__(self, input_csv, output_excel="output.xlsx"):
        self.input_csv = input_csv
        self.output_excel = output_excel
        self.df = None
        self.df_long = None

    def load_data(self):
        self.df = pd.read_csv(self.input_csv)
        self.df = self.df.drop(["Section", "SIS Login ID"], axis=1)
        self.df = self.df.rename(columns={
            "Enrol No": "Enrolment No",
            "SIS User ID": "Student Number"
        })
        return self

    def unpivot(self):
        id_vars = ["Enrolment No", "Student Number"]
        score_cols = [c for c in self.df.columns if c not in id_vars]

        self.df_long = self.df.melt(
            id_vars=id_vars,
            value_vars=score_cols,
            var_name="Module Code",
            value_name="Outcome"
        )
        self.df_long["Finish Date"] = "DD/MM/YYYY"
        return self

    @staticmethod
    def convert_outcome(x):
        try:
            val = float(x)
            if val == 100:
                return "Competency achieved/Pass"
            return "Competency not achieved/Fail"
        except:
            pass

        if pd.isna(x):
            return "Record not found"
        if isinstance(x, str) and x.lower() == "ex":
            return "Credit Transfer/national recognition"

        return "Invalid input"

    def apply_outcomes(self):
        self.df_long["Outcome"] = self.df_long["Outcome"].apply(self.convert_outcome)
        return self

    # def export(self):
    #     with pd.ExcelWriter(self.output_excel, engine="openpyxl") as writer:
    #         self.df.to_excel(writer, "Original", index=False)
    #         self.df_long.to_excel(writer, "FINAL RESULTS", index=False)

    #     self.apply_colours()
    #     print(f"✅ File saved with formatting: {self.output_excel}")
    def export(self, finish_date=None):
        # Insert Finish Date (dd/mm/yyyy) into df_long
        if finish_date is not None:
            self.df_long["Finish Date"] = finish_date

        # Save Excel file
        with pd.ExcelWriter(self.output_excel, engine="openpyxl") as writer:
            self.df.to_excel(writer, "Original", index=False)
            self.df_long.to_excel(writer, "FINAL RESULTS", index=False)

        # Apply colours like before
        self.apply_colours()
        print(f"✅ File saved with formatting: {self.output_excel}")

    def apply_colours(self):
        wb = load_workbook(self.output_excel)
        ws = wb["FINAL RESULTS"]

        green = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        yellow = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        outcome_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(1, col).value == "Outcome":
                outcome_col = col
                break

        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, outcome_col)
            if cell.value == "Competency achieved/Pass":
                cell.fill = green
            elif cell.value == "Credit Transfer/national recognition":
                cell.fill = yellow
            elif cell.value == "Competency not achieved/Fail":
                cell.fill = red
            else:
                cell.fill = gray

        wb.save(self.output_excel)


